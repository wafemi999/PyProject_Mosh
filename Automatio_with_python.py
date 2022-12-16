import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path

'''
# access cell by method 1 cell = sheet['a1']

# access cell by method 2
cell = sheet.cell(1, 1)

# print(cell.value)

# print no of rows:2
print(sheet.max_row)

# accessing each cell on the 3rd column
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)
'''


def process_workbook(filename):
    # load workbook
    book = xl.load_workbook(filename)
    # load sheet
    sheet = book['sheet1']
    # print(dir(openpyxl.chartsheet))
    for row in range(2, sheet.max_row - 995):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        print(corrected_price_cell.value)

    values = Reference(sheet,
                       min_row=2,
                       max_row=4,
                       min_col=4,
                       max_col=4)
    # adding chart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')
    book.save(filename)


process_workbook("dept_emp.xlsx")
